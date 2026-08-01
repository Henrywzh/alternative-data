"""C&SD Table E705 cross-boundary movement statistics."""

from __future__ import annotations

import io
import re
from typing import Any

import pandas as pd
import requests
from openpyxl import load_workbook

from ..config import (
    CENSTATD_BOUNDARY_PRODUCT_URL,
    CENSTATD_BOUNDARY_REPORT_INDEX_URL,
    DEFAULT_HEADERS,
    DEFAULT_TIMEOUT,
)
from ..storage import save_raw_snapshot

SCHEMA_COLUMNS = [
    "date",
    "month",
    "aircraft_arrivals",
    "aircraft_departures",
    "aircraft_total",
    "ocean_vessels_arrivals",
    "ocean_vessels_departures",
    "ocean_vessels_arrival_thousand_nt",
    "ocean_vessels_departure_thousand_nt",
    "river_vessels_prc_arrivals",
    "river_vessels_prc_departures",
    "river_vessels_prc_arrival_thousand_nt",
    "river_vessels_prc_departure_thousand_nt",
    "river_vessels_macao_arrivals",
    "river_vessels_macao_departures",
    "river_vessels_macao_arrival_thousand_nt",
    "river_vessels_macao_departure_thousand_nt",
    "cargo_vessels_arrivals",
    "cargo_vessels_departures",
    "cargo_vessels_arrival_thousand_nt",
    "cargo_vessels_departure_thousand_nt",
    "goods_vehicles_arrivals",
    "goods_vehicles_departures",
    "goods_vehicles_total",
    "passenger_vehicles_arrivals",
    "passenger_vehicles_departures",
    "passenger_vehicles_total",
    "passenger_trains_arrivals",
    "passenger_trains_departures",
    "passenger_trains_total",
    "is_estimate",
]

# The workbook's merged headers leave spacer columns between each measure.
VALUE_COLUMNS = {
    "aircraft_arrivals": 3,
    "aircraft_departures": 5,
    "aircraft_total": 7,
    "ocean_vessels_arrivals": 10,
    "ocean_vessels_arrival_thousand_nt": 11,
    "ocean_vessels_departures": 13,
    "ocean_vessels_departure_thousand_nt": 14,
    "river_vessels_prc_arrivals": 16,
    "river_vessels_prc_arrival_thousand_nt": 17,
    "river_vessels_prc_departures": 19,
    "river_vessels_prc_departure_thousand_nt": 20,
    "river_vessels_macao_arrivals": 22,
    "river_vessels_macao_arrival_thousand_nt": 23,
    "river_vessels_macao_departures": 25,
    "river_vessels_macao_departure_thousand_nt": 26,
    "cargo_vessels_arrivals": 28,
    "cargo_vessels_arrival_thousand_nt": 29,
    "cargo_vessels_departures": 31,
    "cargo_vessels_departure_thousand_nt": 32,
    "goods_vehicles_arrivals": 34,
    "goods_vehicles_departures": 36,
    "goods_vehicles_total": 38,
    "passenger_vehicles_arrivals": 41,
    "passenger_vehicles_departures": 43,
    "passenger_vehicles_total": 45,
    "passenger_trains_arrivals": 48,
    "passenger_trains_departures": 50,
    "passenger_trains_total": 52,
}


def _number(value: Any) -> tuple[float | None, bool]:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None, False
    text = str(value).strip()
    if not text or text.upper() in {"N.A.", "N/A", "-"}:
        return None, False
    estimate = "#" in text
    cleaned = re.sub(r"[^0-9.\-]", "", text)
    if not cleaned:
        return None, estimate
    try:
        return float(cleaned), estimate
    except ValueError as exc:
        raise ValueError(f"C&SD Table E705 has an unparseable value: {value!r}") from exc


def _integer(value: Any) -> int | None:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    try:
        number = float(str(value).strip())
    except (TypeError, ValueError):
        return None
    return int(number) if number.is_integer() else None


def parse_boundary_movements_workbook(payload: bytes) -> pd.DataFrame:
    workbook = load_workbook(io.BytesIO(payload), data_only=True, read_only=True)
    if "E705" not in workbook.sheetnames:
        raise ValueError(f"C&SD Table E705 workbook has unexpected sheets: {workbook.sheetnames}")
    worksheet = workbook["E705"]
    records: list[dict[str, Any]] = []
    current_year: int | None = None
    for row in worksheet.iter_rows(values_only=True):
        first = row[0] if len(row) > 0 else None
        month_value = row[1] if len(row) > 1 else None
        year = _integer(first)
        if year is not None and 1900 <= year <= 2200:
            current_year = year
        if current_year is None or month_value is None:
            continue
        month = _integer(month_value)
        if month is None or not 1 <= month <= 12:
            continue
        record: dict[str, Any] = {
            "date": f"{current_year}-{month:02d}",
            "month": f"{current_year}-{month:02d}",
            "is_estimate": False,
        }
        for field, column in VALUE_COLUMNS.items():
            value, estimate = _number(row[column] if column < len(row) else None)
            record[field] = value
            record["is_estimate"] = record["is_estimate"] or estimate
        records.append(record)
    result = pd.DataFrame(records, columns=SCHEMA_COLUMNS).drop_duplicates("month")
    result = result.sort_values("date").reset_index(drop=True)
    if result.empty:
        raise ValueError("C&SD Table E705 contained no monthly rows")
    return result


def _latest_product() -> tuple[str, str]:
    response = requests.get(
        CENSTATD_BOUNDARY_REPORT_INDEX_URL,
        headers=DEFAULT_HEADERS,
        timeout=max(DEFAULT_TIMEOUT, 30),
    )
    response.raise_for_status()
    data = response.json()
    for product in data.get("productIndex", []):
        if product.get("Product_Code") == "D7000005" and product.get("en_file"):
            return str(product["en_file"]), str(product.get("Issue") or "")
    raise ValueError("C&SD subject 340 report index did not advertise D7000005")


def fetch_censtatd_boundary_movements() -> pd.DataFrame:
    filename, issue = _latest_product()
    url = CENSTATD_BOUNDARY_PRODUCT_URL.format(filename=filename)
    response = requests.get(url, headers=DEFAULT_HEADERS, timeout=max(DEFAULT_TIMEOUT, 30))
    response.raise_for_status()
    result = parse_boundary_movements_workbook(response.content)
    raw_path = save_raw_snapshot(
        "censtatd_boundary_movements",
        response.content,
        file_ext="xlsx",
        source_url=url,
    )
    result.attrs["raw_snapshot"] = str(raw_path)
    result.attrs["source_url"] = url
    result.attrs["source_issue"] = issue
    return result
