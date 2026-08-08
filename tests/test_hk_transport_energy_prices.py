from __future__ import annotations

import io

import pandas as pd
from openpyxl import Workbook

from src.hk_transport.sources.energy_prices import (
    ENERGY_PRICE_COLUMNS,
    parse_eia_spot_price_workbook,
)


def _workbook_bytes() -> bytes:
    workbook = Workbook()
    contents = workbook.active
    contents.title = "Contents"
    contents.append([None, "Workbook Contents"])
    contents.append([None, "Release Date:", "08/05/2026"])

    crude = workbook.create_sheet("Data 1")
    crude.append(["Back to Contents", "Data 1: Crude Oil", None])
    crude.append(["Sourcekey", "RWTC", "RBRTE"])
    crude.append(["Date", "WTI", "Brent"])
    crude.append(["2026-07-31", 84.51, 91.63])

    jet = workbook.create_sheet("Data 6")
    jet.append(["Back to Contents", "Data 6: Kerosene-Type Jet Fuel"])
    jet.append(["Sourcekey", "EER_EPJK_PF4_RGC_DPG"])
    jet.append(["Date", "Jet fuel"])
    jet.append(["2026-07-31", 3.736])

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_parse_eia_spot_price_workbook_preserves_vintage_and_units() -> None:
    result = parse_eia_spot_price_workbook(
        _workbook_bytes(),
        frequency="weekly",
        source_url="https://example.test/eia.xls",
        retrieved_at="2026-08-06T00:00:00+00:00",
    )

    assert list(result.columns) == ENERGY_PRICE_COLUMNS
    assert set(result["series_id"]) == {"RWTC", "RBRTE", "EER_EPJK_PF4_RGC_DPG"}
    assert set(result["unit"]) == {"USD per barrel", "USD per gallon"}
    assert set(result["source_release_date"]) == {"2026-08-05"}
    assert set(result["retrieved_at"]) == {"2026-08-06T00:00:00+00:00"}
    assert result.loc[result["series_id"].eq("EER_EPJK_PF4_RGC_DPG"), "value"].item() == 3.736
